import math
import numpy as np
import heapq
import matplotlib.pyplot as plt
import matplotlib.ticker as plticker
import csv
import ast
import sys
import time
import copy
import warnings
import glob
import os
import glob
import pandas as pd
import openpyxl as openpyxl
warnings.filterwarnings("ignore")
import utils
from tkinter import *
import config





class MDR():


    def _compute_makespan(routes):
        return max([len(route) for route in routes])

    ##############################################################################
    # heuristic function for heuristic_interrupt path scoring
    ##############################################################################
    '''
    search_node contains position of all agent, allowed abnormal moves, current time step
    routes contains the planned routes for all agents [ including the path they passed so far ]
    '''
    def f_interrupt(self, search_node, routes):

        cost_without_interrupts = MDR._compute_makespan(routes)

        # An overestimate of the cost added by the abnormal moves
        remaining_abnormal_moves = search_node.k
        num_of_agents = len(search_node.pos)

        # The idea is that the maximal addition for each abnormal move is that it will delay all agents by one time step.
        return -1*(cost_without_interrupts + num_of_agents * remaining_abnormal_moves)  # TODO: CHANGE 2

        '''
         route = list containing the planned route for each agent 
        '''

    def is_goal(self, node, route):

        for x in range(0, len(node.pos)):
            if node.pos[x] is not route[x][-1][0] and node.pos[x] is not None:
                return False
        return True

    def apply(self, action, node, step, route_, robust_factor):
        #route = copy.deepcopy(route_)
        routes = [route for route in route_]
        tmp_node = list()
        pos = list()

        if step < len(routes[0]):
            # If action of Agent 0 is according the plan - don't decrease the no. of remained error
            if routes[0][step][0] == (node.pos[0][0] + action[0], node.pos[0][1] + action[1]):
                pos.append(routes[0][step][0])
                for w in range(1, len(routes)):
                    if step < len(routes[w]):
                        pos.append(routes[w][step][0])
                    else:
                        pos.append(routes[w][-1][0])
                tmp_node = MDR.Node(pos, node.k, step, node, route=routes)
                return tmp_node, routes

            # If action of Agent 0 is not according the plan - decrease the no. of remained error and sent to re-calculate new routes
            else:
                if robust_factor > 1:
                    robust_factor = robust_factor - 1
                else:
                    robust_factor = 0
                tmp_node, routes = self.calc_new_routes(action, node, step, routes, robust_factor)

        else:
            return None, None

        return tmp_node, routes

    def calc_new_routes(self, action, node, step, route_, robust_factor):

        routes = [route for route in route_]
        tmp_node = list()
        pos = list()

        # set the no. of agent that making the move for interruptions
        adversarial_agent = 0

        tmp_node.append((node.pos[0][0] + action[0], node.pos[0][1] + action[1]))

        ###########################################################################
        # Send Agent - 0 to A* (grid, start = next new step, goal, agent no., step)
        ###########################################################################

        # If there are more steps, (i.e. it is not at the end of it's route) for agent 0,
        # so run A* for Agent - 0
        if step < len(routes[0]):
            new_route = []
            for h in range(0, step):
                new_route.append(routes[adversarial_agent][h])

            start = ((tmp_node[0]), step)
            goal = routes[adversarial_agent][-1][0]
            if step >= 1:
                routes[adversarial_agent] = astar(node, self, start, goal, adversarial_agent, routes, step, robust_factor, self.db_heuristic)
                for h in routes[adversarial_agent]:
                    new_node = ((h[0]), new_route[-1][1] + 1)
                    new_route.append(new_node)
                routes[adversarial_agent] = new_route
            else:
                routes[adversarial_agent] = astar(node, self, start, goal, adversarial_agent, routes, step, robust_factor, self.db_heuristic)
        else:
            new_node = routes[adversarial_agent][-1][0]
            routes[adversarial_agent].append(new_node)

        ##########################################################
        # Build the new Node after agent - 0 build its new route.
        ##########################################################

        pos.append(tmp_node[0])
        for x in range(1, len(routes)):
            if step < len(routes[x]):
                pos.append(routes[x][step][0])
            else:
                pos.append(routes[x][-1][0])
        tmp_node = MDR.Node(pos, node.k - 1, step, node, route=routes)

        return tmp_node, routes

    def search_for_interrupt_plan(self, grid, route, k, robust_factor):
        total_expanded_nodes = 0
        actions = [(0, 1), (0, -1), (1, 0), (-1, 0), (1, 1), (1, -1), (-1, 1), (-1, -1), (0, 0)]

        open = []
        pos = list()
        step = 0
        self.grid = grid

        # Set starting positions for all agents into a node
        num_of_agents = len(route)
        for x in range(0, num_of_agents):
            pos.append(route[x][0][0])

        print("Pre-computing shortest paths from goals...", end='')
        goal_to_db = dict()
        for x in range(0, num_of_agents):
            agent_goal = route[x][-1][0]
            operators_and_costs = [(0, 1,1), (0, -1,1), (1, 0,1), (-1, 0,1), (1, 1,1), (1, -1,1), (-1, 1,1), (-1, -1,1) , (0, 0,1)]
            goal_to_db[agent_goal] = utils.dijkstra(self.grid, agent_goal, operators_and_costs)

        self.db_heuristic = DatabaseHeuristic(goal_to_db)
        print("Done")


        start = MDR.Node(pos=pos, k=k, step=step, parent=None, route=route)
        f_score = self.f_interrupt(start, route)
        fscore = {start: f_score}
        heapq.heappush(open, (fscore[start], start))

        goals = list()
        best_goal_value = MDR._compute_makespan(route) # TODO Change

        while (len(open) > 0):
            (f, node) = heapq.heappop(open)

            # count total expanded nodes
            total_expanded_nodes = total_expanded_nodes + 1
            if total_expanded_nodes % 10000 == 0:
                print("Total expanded nodes so far = %d, max f in OPEN = %d" % (total_expanded_nodes, f))
            step = node.step + 1

            # No need to expand a node that cannot improve the best goal
            node_value = -f
            if node_value < best_goal_value: # TODO: Change 4
                continue
            if self.is_goal(node, node.route):
                continue

            # print("%d" % (len(goals)))

            for action in actions:

                if step < len(route[0]):
                    planned_radians = math.atan2(route[0][step][0][1] - node.pos[0][1],
                                                 route[0][step][0][0] - node.pos[0][0])
                    planned_degrees = math.degrees(planned_radians)
                else:
                    planned_radians = math.atan2(route[0][-1][0][1] - node.pos[0][1],
                                                 route[0][-1][0][0] - node.pos[0][0])
                    planned_degrees = math.degrees(planned_radians)
                # print(planned_degrees)

                next_step_radians = math.atan2((node.pos[0][1] + action[1]) - node.pos[0][1],
                                               (node.pos[0][0] + action[0]) - node.pos[0][0])
                next_step_degrees = math.degrees(next_step_radians)

                anglediff = (next_step_degrees - planned_degrees + 180 + 360) % 360 - 180

                if (anglediff > 90 or anglediff < -90):
                    continue

                if node.k == 0:
                    # no more err to use
                    pos = list()
                    for x in range(0, len(node.route)):
                        if step < len(node.route[x]):
                            pos.append(node.route[x][step][0])
                        else:
                            # TODO: Try to add none instead the last node and verify the code
                            pos.append(node.route[x][-1][0])
                    new_node = MDR.Node(pos, 0, step, node, node.route)
                    if (self.is_goal(new_node, new_node.route)):
                        goals.append(new_node)

                        # find the total cost of the current route
                        cost_new_node = MDR._compute_makespan(new_node.route) # TODO Change

                        if cost_new_node > best_goal_value:
                            best_goal_value = cost_new_node

                    else:
                        f_score = self.f_interrupt(new_node, new_node.route) ## RONI: NEW LINE
                        fscore[new_node] = f_score
                        heapq.heappush(open, (fscore[new_node], new_node))

                    break

                # if action is legal (not collide in wall)
                if self.grid[node.pos[0][0] + action[0]][node.pos[0][1] + action[1]] == 1:
                    # array bound 1 on map
                    continue

                new_node, _ = self.apply(action, node, step, node.route, robust_factor)
                if new_node == None:
                    continue
                if (self.is_goal(new_node, new_node.route)):
                    goals.append(new_node)

                    # find the total cost of the current route
                    cost_new_node = MDR._compute_makespan(new_node.route)

                    if cost_new_node > best_goal_value:
                        best_goal_value = cost_new_node

                else:
                    f_score = self.f_interrupt(new_node, new_node.route)
                    fscore[new_node] = f_score # + node.step  # + (np.random.random()/1) # TODO: CHANGE 3
                    heapq.heappush(open, (fscore[new_node], new_node))

        # After all goals have been found
        max_step = 0
        best_goal = None
        for goal in goals:
            if goal.step > max_step:
                max_step = goal.step
                best_goal = goal

        # Reconstruct the solution
        solution = [best_goal]
        parent = best_goal.parent

        while parent is not None:
            solution.append(parent)
            parent = parent.parent

        solution = list(reversed(solution))

        return (solution, route, total_expanded_nodes)

    class Node:

        def __init__(self, pos, k, step, parent, route=None):
            self.pos = pos
            self.k = k
            self.step = step
            self.parent = parent
            self.route = route

        def weight(self):
            next_step_radians = math.atan2((self.route[0][-1][0][1]) - self.pos[0][1],
                                           (self.route[0][-1][0][0]) - self.pos[0][0])
            w = math.degrees(next_step_radians)

            return w

        def __lt__(self, other):
            return self.weight() < other.weight()



##############################################################################
# heuristic function for A* path scoring
##############################################################################
class Heuristic:
    def value(self, a, b): raise NotImplementedError

class ManhattanDistance(Heuristic):
    def value(self, a, b):
        distance = (abs(b[0] - a[0]) + abs(b[1] - a[1]))
        return distance

'''
    This heuristic accepts a database that maps a pair of nodes to the distance between them. 
'''
class DatabaseHeuristic(Heuristic):
    def __init__(self, goal_to_db):
        self.goal_to_db = goal_to_db

    def value(self, a, b):
        return self.goal_to_db[b][a]


def in_start_state(current, route, stp, y):
    is_in_start = False
    for check_start in range(0, stp):
        if route[y][check_start][0] == current[0]:
            is_in_start = True
        else:
            return False
    return is_in_start

##############################################################################
# path finding function (not to be confused with the A* run by the MDR algorithm)
##############################################################################
def astar(current_node, array, start, goal, adversarial_agent, route, step, robust_factor, h=ManhattanDistance):
    neighbors = [(0, 1), (0, -1), (1, 0), (-1, 0), (1, 1), (1, -1), (-1, 1), (-1, -1), (0, 0)]

    start_node = (start)
    close_set = set()
    came_from = {start_node: 0}
    gscore = {start_node: 0}
    fscore = {start_node: h.value(start[0], goal)}
    oheap = []
    heapq.heappush(oheap, (fscore[start_node], start_node))
    is_collision = 0
    data = []

    while oheap:
        current = heapq.heappop(oheap)[1]
        if current[0] == goal:
            while current in came_from:
                data.append((current[0], current[1]))
                current = came_from[current]
            data = list(reversed(data))
            return data

        close_set.add(current)

        action_cost = 1  # TODO: Replace this in the future to support different action costs

        for op in neighbors:
            (i, j) = op
            neighbor = ((current[0][0] + i, current[0][1] + j), current[1] + action_cost)
            neighbor_is_valid = 0
            tentative_g_score = gscore[current] + action_cost

            # Get valid steps (radius)
            valid_radius = utils.get_MDR_radius(array.grid, neighbor, route, adversarial_agent, step, robust_factor)

            if not valid_radius:
                continue

            if neighbor in close_set and tentative_g_score >= gscore.get(neighbor[0], 0):
                continue

            if tentative_g_score < gscore.get(neighbor[0], 0) or neighbor not in [i[1] for i in oheap]:
                came_from[neighbor] = (current)
                gscore[neighbor] = tentative_g_score
                fscore[neighbor] = tentative_g_score + h.value(neighbor[0], goal)
                heapq.heappush(oheap, (fscore[neighbor], neighbor))

    return data


###################################################################################
# INPUT UI
###################################################################################
global map_file_name
global data_folder
global DS_Budget

def show_entry_fields():
    print("Room File: %s\nData Folder: %s\nMDR Factor: %s\nRobust Factor: %s" % (e1.get(), e2.get(), e3.get(), e4.get()))
master = Tk()
Label(master, text="Room File").grid(row=0)
Label(master, text="Data Folder").grid(row=1)
Label(master, text="Damage Steps Budget").grid(row=2)

e1 = Entry(master)
e2 = Entry(master)
e3 = Entry(master)

print(config.var_a)

e1.insert(10, config.var_a)
e2.insert(10, config.var_b)
e3.insert(10, config.var_c)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)
e3.grid(row=2, column=1)

Button(master, text='Run', command=master.quit).grid(row=7, column=0, sticky=W, pady=4)

mainloop()

map_file_name = e1.get()
data_folder = e2.get()
DS_Budget = int(e3.get())


'''
The entry point to running this algorithm
'''
def main():
    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute.
    sheet = wb.active

    # load the map file into numpy array
    with open(map_file_name, 'rt') as infile:
        grid1 = np.array([list(line.strip()) for line in infile.readlines()])
    print('Grid shape', grid1.shape)

    grid1[grid1 == '@'] = 1  # object on the map
    grid1[grid1 == 'T'] = 1  # object on the map
    grid1[grid1 == '.'] = 0  # free on map

    grid = np.array(grid1.astype(np.int))

    # Plot the map
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.imshow(grid, cmap=plt.cm.ocean)

    path = data_folder + '\\'

    extension = 'csv'
    os.chdir(path)
    Sub_folders = next(os.walk('.'))[1]

    # Run over all sub folders
    for sub_folder in range(0, len(Sub_folders)):

        robust_factor = int(Sub_folders[sub_folder])

        path = str(Sub_folders[sub_folder]) + '\\'
        os.chdir(path)

        files = [i for i in glob.glob('*.{}'.format(extension))]

        counter = 1

        res = ['File Name', 'Leader No.', 'DS Budget', 'No. of Agents', 'Run Time',
               'Org. MS', 'New MS']
        # writing to the specified cell
        for len_of_row in range(1, len(res)+1):
            sheet.cell(row=counter, column=len_of_row).value = res[len_of_row-1]
        wb.save('results.xlsx')
        counter = counter+1

        # Run over all MAPF plan files in the folder
        for file in files:
            tmp_main_route = []
            main_route = []

            # Reading csv file
            with open(file, 'rt') as csvfile:

                # creating a csv reader object
                csvreader = csv.reader(csvfile)

                # extracting each data row one by one
                for row in csvreader:
                    tmp_route = []
                    for y in range(0, len(row)):
                        if row[y]:
                            tmp_route.append(ast.literal_eval(row[y]))
                    tmp_main_route.append(tmp_route)

                    # Remove the first row
                    main_route = tmp_main_route[:]
                    main_route.remove(tmp_main_route[0])

            org_MS = max([len(org_route) for org_route in main_route])

            # Run only a0 as the leader as the plan is robust for a1 as adversarial agent
            # Change here to run over all agents as leaders
            for current_agent in range(0, 1): #len(main_route)):

                # LOOP for No. of damage steps between 1-MDR Factor errors
                for no_of_DS in range(0, DS_Budget + 1):

                    start_time = time.time()

                    print('######################################################################')
                    print('Run Number - ', counter - 1, file)
                    print('######################################################################')
                    mdr_solver = MDR()
                    solution, route, total_expanded_nodes = mdr_solver.search_for_interrupt_plan(grid, main_route, no_of_DS, robust_factor)

                    elapsed_time = time.time() - start_time

                    # Write the new MDR path to sub folder
                    mdr_route = []
                    path = str('MDR_New_Routes') + '\\'

                    if not os.path.exists(path):
                        os.makedirs(path)

                    os.chdir(path)
                    for step in range(0, len(solution)):
                        mdr_route.append(solution[step].pos)


                    tmp_mdr_route = [*zip(*mdr_route)]
                    # Write to Excel
                    df_res = pd.DataFrame(tmp_mdr_route)

                    file_name = []
                    file_name.append(str(file))
                    file_name.append('No of DS - ')
                    file_name.append(str(no_of_DS))
                    file_name.append('.xlsx')
                    file_n = ' '.join(file_name)
                    file_n.replace(" ", "")

                    writer_xl = pd.ExcelWriter(file_n, engine='openpyxl')
                    df_res.to_excel(writer_xl, sheet_name='Agents', index=False)
                    writer_xl.save()
                    writer_xl.close()
                    path = '..'
                    os.chdir(path)
                    # End of write to file

                    # Find the maximum of the new makespan (New_MD) not include with a0,
                    # i.e. how much damage can a0 make to the swarm but not to itself.


                    new_MS = 0
                    for find_new_MS in range(1, len(tmp_mdr_route)):
                        tmp_new_MS = 0
                        # Set goal for each route
                        mdr_goal = tmp_mdr_route[find_new_MS][len(tmp_mdr_route[find_new_MS]) - 1]
                        for i in range(0, len(tmp_mdr_route[find_new_MS])):
                            if tmp_mdr_route[find_new_MS][i] == mdr_goal:
                                tmp_new_MS = i + 1
                                break
                        if tmp_new_MS > new_MS:
                            new_MS = tmp_new_MS

                    print('Total runtime - ' + str(elapsed_time))
                    print('Robust Level Org Plan - ', robust_factor)
                    print('Number of Errors - ', no_of_DS)
                    print('Original make span - ', org_MS)
                    print('New make span - ', new_MS)


                    res = [file, current_agent, no_of_DS, len(route), elapsed_time, org_MS, new_MS]

                    # writing to the specified cell
                    for len_of_row in range(1, len(res)+1):
                        sheet.cell(row=counter, column=len_of_row).value = res[len_of_row - 1]
                    counter = counter + 1

            wb.save('results.xlsx')

        path = '..'
        os.chdir(path)


'''
Standard python code for making sure what should run when this module is run as the main script 
as oppose to when this code is imported. I'm adding this to allow unit testing.
'''
if __name__ == '__main__':
    main()